#!/usr/bin/env python3
"""Imposta la sezione bollettino nel template Polline_Template_Settimanale.xlsx.

Pre-popola tutte le specie di SOGLIE_MAPPING con formule Excel che referenziano
la tabella dati grezzi, imposta la formattazione condizionale per i colori e
le intestazioni con date dinamiche (da J3).

Eseguire manualmente dopo aver modificato SOGLIE_MAPPING o le soglie di
concentrazione. Non e' necessario eseguire questo script ad ogni sessione.
"""

import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "codice"))

import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

from polline_counter import (
    BOLL_START_ROW,
    CODICI_SPECIE,
    SOGLIE_MAPPING,
    carica_soglie,
    codice_to_row,
    giorno_to_col,
)

TEMPLATE_PATH = Path(__file__).parent.parent / "codice" / "Polline_Template_Settimanale.xlsx"

# Layout colonne bollettino (identico a genera_bollettino)
T1_START = 4   # D
T2_START = 16  # P

# Stili
_thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_giallo     = PatternFill("solid", fgColor="FFFF00")
_blu        = PatternFill("solid", fgColor="4472C4")
_grigio     = PatternFill("solid", fgColor="808080")
_f_nero     = Font(color="000000", size=10)
_f_nero_b   = Font(color="000000", bold=True, size=11)
_f_bianco_b = Font(color="FFFFFF", bold=True, size=11)
_center     = Alignment(horizontal="center", vertical="center")
_center_w   = Alignment(horizontal="center", wrap_text=True)

_FILL_ASSENTE = PatternFill("solid", fgColor="00B050")
_FILL_BASSA   = PatternFill("solid", fgColor="FFD966")
_FILL_MEDIA   = PatternFill("solid", fgColor="F4B084")
_FILL_ALTA    = PatternFill("solid", fgColor="FF0000")


def _dxf(fill):
    return DifferentialStyle(fill=fill)


def setup_bollettino(ws, soglie):
    no_fill   = PatternFill(fill_type=None)
    no_border = Border()

    # Pulisci area bollettino (righe 72-120, colonne D-Y)
    for r in range(72, 121):
        for c in range(T1_START, 26):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = no_fill
            cell.border = no_border
            cell.font = Font()
            cell.alignment = Alignment()
            cell.number_format = "General"

    # Riga 72: separatore visivo grigio
    for c in range(T1_START, 26):
        ws.cell(row=72, column=c).fill = _grigio

    # Riga 73: titolo con mese e anno dinamici da J3
    title_cell = ws.cell(
        row=BOLL_START_ROW, column=T1_START,
        value='=IFERROR("BOLLETTINO POLLINICO - "&PROPER(TEXT($J$3,"MMMM"))&" "&TEXT($J$3,"YYYY"),"BOLLETTINO POLLINICO")',
    )
    title_cell.font = Font(bold=True, size=14)

    # Riga 75 (BOLL_START_ROW+2): intestazioni colonne
    # Intestazioni T1 (sfondo giallo)
    headers_t1 = [
        (T1_START, "Famiglia/Specie"),
        *[
            (T1_START + 1 + g, f'=IFERROR(PROPER(TEXT($J$3+{g},"DDDD"))&" "&DAY($J$3+{g}),"Giorno {g+1}")')
            for g in range(7)
        ],
        (T1_START + 8, "Media (p/m3)"),
    ]
    for col, val in headers_t1:
        cell = ws.cell(row=BOLL_START_ROW + 2, column=col, value=val)
        cell.font = _f_nero_b
        cell.fill = _giallo
        cell.alignment = _center_w
        cell.border = _thin

    # Intestazioni T2 (sfondo blu)
    headers_t2 = [
        (T2_START, "Famiglia/Specie"),
        *[
            (T2_START + 1 + g, f'=IFERROR(PROPER(TEXT($J$3+{g},"DDDD"))&" "&DAY($J$3+{g}),"Giorno {g+1}")')
            for g in range(7)
        ],
        (T2_START + 8, "Media (p/m3)"),
        (T2_START + 9, "Tendenza"),
    ]
    for col, val in headers_t2:
        cell = ws.cell(row=BOLL_START_ROW + 2, column=col, value=val)
        cell.font = _f_bianco_b
        cell.fill = _blu
        cell.alignment = _center_w
        cell.border = _thin

    # Costruisce la lista specie nell'ordine di SOGLIE_MAPPING
    specie_list = []
    for codice, famiglia_soglia in SOGLIE_MAPPING.items():
        row_riep = codice_to_row(codice)
        if row_riep is None:
            continue
        nome = CODICI_SPECIE.get(codice, famiglia_soglia)
        if codice in ("48", "50"):
            nome = f"Spore fungine di {nome}"
        soglia_tuple = soglie.get(famiglia_soglia, (0.9, 19.9, 39.9))
        specie_list.append((nome, row_riep, soglia_tuple))

    riga_dati_start = BOLL_START_ROW + 3
    cf_priority = 1  # priorita' globale unica per tutte le regole CF del foglio

    for i, (nome, row_riep, soglia_tuple) in enumerate(specie_list):
        riga = riga_dati_start + i
        max_assente, max_bassa, max_media = soglia_tuple

        # --- Tabella T1 (sinistra, senza colore) ---
        c_nome1 = ws.cell(row=riga, column=T1_START, value=nome)
        c_nome1.font = _f_nero
        c_nome1.border = _thin

        for g in range(7):
            col_raw = get_column_letter(giorno_to_col(g + 1))
            cell = ws.cell(
                row=riga, column=T1_START + 1 + g,
                value=f"=IF({col_raw}{row_riep}>0,ROUND({col_raw}{row_riep}*$Q$3,1),0)",
            )
            cell.font = _f_nero
            cell.alignment = _center
            cell.border = _thin
            cell.number_format = "0.0"

        c_media1 = ws.cell(
            row=riga, column=T1_START + 8,
            value=f"=ROUND(SUM(G{row_riep}:M{row_riep})*$Q$3/7,1)",
        )
        c_media1.font = _f_nero
        c_media1.alignment = _center
        c_media1.border = _thin
        c_media1.number_format = "0.0"

        # --- Tabella T2 (destra, colorata) ---
        c_nome2 = ws.cell(row=riga, column=T2_START, value=nome)
        c_nome2.font = _f_nero
        c_nome2.border = _thin

        for g in range(7):
            col_raw = get_column_letter(giorno_to_col(g + 1))
            cell = ws.cell(
                row=riga, column=T2_START + 1 + g,
                value=f"=IF({col_raw}{row_riep}>0,ROUND({col_raw}{row_riep}*$Q$3,1),0)",
            )
            cell.font = _f_nero
            cell.alignment = _center
            cell.border = _thin
            cell.number_format = "0.0"

        c_media2 = ws.cell(
            row=riga, column=T2_START + 8,
            value=f"=ROUND(SUM(G{row_riep}:M{row_riep})*$Q$3/7,1)",
        )
        c_media2.font = _f_nero
        c_media2.alignment = _center
        c_media2.border = _thin
        c_media2.number_format = "0.0"

        c_tend = ws.cell(row=riga, column=T2_START + 9)
        c_tend.border = _thin
        c_tend.alignment = _center

        ws.row_dimensions[riga].height = 15

        # Formattazione condizionale T2 (Q:X = 7 giorni + media)
        # Le regole sono non-sovrapposte grazie a stopIfTrue=True in ordine crescente
        rng = f"Q{riga}:X{riga}"
        for formula_val, dxf in [
            (str(max_assente), _dxf(_FILL_ASSENTE)),
            (str(max_bassa),   _dxf(_FILL_BASSA)),
            (str(max_media),   _dxf(_FILL_MEDIA)),
        ]:
            ws.conditional_formatting.add(rng, Rule(
                type="cellIs", operator="lessThanOrEqual",
                formula=[formula_val], dxf=dxf,
                stopIfTrue=True, priority=cf_priority,
            ))
            cf_priority += 1
        ws.conditional_formatting.add(rng, Rule(
            type="cellIs", operator="greaterThan",
            formula=[str(max_media)], dxf=_dxf(_FILL_ALTA),
            priority=cf_priority,
        ))
        cf_priority += 1

    # Legenda
    riga_legenda = riga_dati_start + len(specie_list) + 1
    ws.cell(row=riga_legenda, column=T1_START,
            value="Concentrazioni di riferimento").font = Font(bold=True)
    for j, (etichetta, fill) in enumerate([
        ("Assente", _FILL_ASSENTE), ("Bassa", _FILL_BASSA),
        ("Media", _FILL_MEDIA), ("Alta", _FILL_ALTA),
    ]):
        cell = ws.cell(row=riga_legenda, column=T1_START + 1 + j, value=etichetta)
        cell.fill = fill
        cell.alignment = _center
        cell.font = Font(bold=True)
        cell.border = _thin

    print(f"  Bollettino: {len(specie_list)} specie, righe {riga_dati_start}-{riga_dati_start + len(specie_list) - 1}")
    print(f"  Formattazione condizionale: {cf_priority - 1} regole")


def main():
    backup = TEMPLATE_PATH.with_name(TEMPLATE_PATH.stem + "_bak.xlsx")
    shutil.copy2(TEMPLATE_PATH, backup)
    print(f"Backup: {backup.name}")

    soglie = carica_soglie()
    if soglie is None:
        print("ERRORE: impossibile caricare le soglie da concentrazioni_polliniche.xlsx")
        sys.exit(1)
    print(f"Soglie caricate: {len(soglie)} famiglie")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["riepilogo_settimana"]

    # J3 e K3 devono essere date Excel (non stringhe) per le formule di intestazione
    ws["J3"].number_format = "DD-MM-YYYY"
    ws["K3"].number_format = "DD-MM-YYYY"

    setup_bollettino(ws, soglie)

    wb.save(TEMPLATE_PATH)
    print(f"Template salvato: {TEMPLATE_PATH.name}")


if __name__ == "__main__":
    main()
