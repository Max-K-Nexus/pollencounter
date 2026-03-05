"""
applica_formattazione.py — script one-shot.

Applica la formattazione visiva di "Tabelle monitoraggi ultima settimana gennaio.xlsx"
al foglio `riepilogo_settimana` di Polline_Template_Settimanale.xlsx
(file principale + copie in linux/ e windows/).

NON tocca: foglio 00_CODICI_SPECIE, foglio dati_grezzi, struttura dati.
"""

import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.formatting import ConditionalFormattingList

# ---------------------------------------------------------------------------
# Configurazione
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent.parent / "codice"
TEMPLATE_NAME = "Polline_Template_Settimanale.xlsx"

TEMPLATE_PRINCIPALE = BASE_DIR / TEMPLATE_NAME
COPIE = [
    BASE_DIR.parent / "windows" / TEMPLATE_NAME,
]

# ---------------------------------------------------------------------------
# Stili
# ---------------------------------------------------------------------------

VERDE = PatternFill(fill_type="solid", fgColor="C5E0B4")
NESSUN_FILL = PatternFill(fill_type=None)

# Colori formattazione condizionale bollettino
CF_VERDE   = PatternFill(fill_type="solid", fgColor="00B050")  # assente
CF_GIALLO  = PatternFill(fill_type="solid", fgColor="FFFF00")  # bassa
CF_ARANCIO = PatternFill(fill_type="solid", fgColor="FF8C00")  # media
CF_ROSSO   = PatternFill(fill_type="solid", fgColor="FF0000")  # alta

thin = Side(border_style="thin", color="000000")
BORDO_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)

FONT_BOLD = Font(bold=True)
FONT_NORMALE = Font(bold=False)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

# Altezza riga in pt (unità openpyxl = pt)
ALTEZZA_RIGHE = 15


# ---------------------------------------------------------------------------
# Bollettino: mappatura famiglie e formattazione condizionale
# ---------------------------------------------------------------------------

# Colonne dati grezzi nel foglio riepilogo_settimana (lun→dom)
_RIEP_GIORNI = ['G', 'H', 'I', 'J', 'K', 'L', 'M']

# Colonne nel bollettino: E=lun, F=mar, ..., K=dom (col 5-11), L=media (col 12)
_BOLL_GIORNI_COLS = range(5, 12)   # colonne E-K
_BOLL_MEDIA_COL   = 12             # colonna L

# (riga_bollettino, righe_riepilogo_da_sommare, assente_max, bassa_max, media_max)
# Solo righe famiglia: contiene la logica di aggregazione delle sottospecie
BOLL_FAMIGLIE = [
    (76,  [6],                 0.9,   19.9,  39.9),  # ACERACEAE
    (77,  [8, 9, 10],          0.5,   15.9,  49.9),  # BETULACEAE (+ Alnus + Betula)
    (80,  [12],                0.0,    4.9,  24.9),  # CHENO-AMAR
    (81,  [13, 14, 15, 16],    0.0,    4.9,  24.9),  # COMPOSITAE (+ sottospecie)
    (85,  [17, 18, 19, 20, 21],0.5,   15.9,  49.9),  # CORYLACEAE (+ sottospecie)
    (90,  [22],                3.9,   29.9,  89.9),  # CUP-TAXACEAE
    (91,  [25, 26, 27, 28],    0.9,   19.9,  39.9),  # FAGACEAE (+ sottospecie)
    (95,  [29],                0.5,    9.9,  29.9),  # GRAMINEAE
    (96,  [36, 37, 38, 39, 40],0.5,    4.9,  24.9),  # OLEACEAE (+ sottospecie)
    (101, [41],                0.9,   14.9,  49.9),  # PINACEAE
    (102, [42],                0.0,    0.4,   1.9),  # PLANTAGINACEAE
    (103, [43],                0.9,   19.9,  39.9),  # PLATANACEAE
    (104, [46, 47, 48],        0.9,   19.9,  39.9),  # SALICACEAE (+ Populus + Salix)
    (107, [50],                0.9,   19.9,  39.9),  # ULMACEAE
    (108, [52],                1.9,   19.9,  69.9),  # URTICACEAE
    (109, [58],                1.9,   19.0, 100.0),  # Alternaria
    (110, [60],              100.0,  499.0,1000.0),  # Cladosporium
]

# (riga_bollettino, assente_max, bassa_max, media_max)
# Tutte le righe (famiglie + sottospecie): usano le soglie della famiglia di appartenenza
BOLL_CF_RIGHE = [
    (76,  0.9,   19.9,  39.9),  # ACERACEAE
    (77,  0.5,   15.9,  49.9),  # BETULACEAE
    (78,  0.5,   15.9,  49.9),  #   Alnus
    (79,  0.5,   15.9,  49.9),  #   Betula
    (80,  0.0,    4.9,  24.9),  # CHENO-AMAR
    (81,  0.0,    4.9,  24.9),  # COMPOSITAE
    (82,  0.0,    4.9,  24.9),  #   Altre compositae
    (83,  0.0,    4.9,  24.9),  #   Ambrosia
    (84,  0.0,    4.9,  24.9),  #   Artemisia
    (85,  0.5,   15.9,  49.9),  # CORYLACEAE
    (86,  0.5,   15.9,  49.9),  #   Carpinus/Ostrya
    (87,  0.5,   15.9,  49.9),  #   Carpinus
    (88,  0.5,   15.9,  49.9),  #   Ostrya carpinifolia
    (89,  0.5,   15.9,  49.9),  #   Corylus avellana
    (90,  3.9,   29.9,  89.9),  # CUP-TAXACEAE
    (91,  0.9,   19.9,  39.9),  # FAGACEAE
    (92,  0.9,   19.9,  39.9),  #   Castanea sativa
    (93,  0.9,   19.9,  39.9),  #   Fagus sylvatica
    (94,  0.9,   19.9,  39.9),  #   Quercus
    (95,  0.5,    9.9,  29.9),  # GRAMINEAE
    (96,  0.5,    4.9,  24.9),  # OLEACEAE
    (97,  0.5,    4.9,  24.9),  #   Altre oleaceae
    (98,  0.5,    4.9,  24.9),  #   Fraxinus
    (99,  0.5,    4.9,  24.9),  #   Ligustrum
    (100, 0.5,    4.9,  24.9),  #   Olea
    (101, 0.9,   14.9,  49.9),  # PINACEAE
    (102, 0.0,    0.4,   1.9),  # PLANTAGINACEAE
    (103, 0.9,   19.9,  39.9),  # PLATANACEAE
    (104, 0.9,   19.9,  39.9),  # SALICACEAE
    (105, 0.9,   19.9,  39.9),  #   Populus
    (106, 0.9,   19.9,  39.9),  #   Salix
    (107, 0.9,   19.9,  39.9),  # ULMACEAE
    (108, 1.9,   19.9,  69.9),  # URTICACEAE
    (109, 1.9,   19.0, 100.0),  # Alternaria
    (110, 100.0, 499.0,1000.0), # Cladosporium
]


def aggiorna_bollettino(ws) -> None:
    """Aggiorna le formule delle righe famiglia nel bollettino (somma famiglia +
    sottospecie) e applica la formattazione condizionale per le soglie.

    Il bollettino è duplicato: lato sinistro (E-L, col 5-12) e lato destro
    (Q-X, col 17-24). Entrambi vengono aggiornati.
    """

    # Colonne lato sinistro: E-K giorni (5-11), L media (12)
    _LX_GIORNI = range(5, 12)
    _LX_MEDIA  = 12

    # Colonne lato destro: Q-W giorni (17-23), X media (24)
    _DX_GIORNI = range(17, 24)
    _DX_MEDIA  = 24

    # Sostituisce l'intera lista CF con una nuova (clear() su _cf_rules non è
    # sufficiente: openpyxl salva anche le regole da una struttura separata)
    ws.conditional_formatting = ConditionalFormattingList()

    for boll_row, riep_rows, assente_max, bassa_max, media_max in BOLL_FAMIGLIE:

        # --- Formule: solo righe famiglia con più righe da sommare ---
        if len(riep_rows) > 1:
            somma_giorni = {
                riep_col: "+".join(f"{riep_col}{r}" for r in riep_rows)
                for riep_col in _RIEP_GIORNI
            }
            sums_settimana = "+".join(f"SUM(G{r}:M{r})" for r in riep_rows)

            # Lato sinistro
            for boll_col, riep_col in zip(_LX_GIORNI, _RIEP_GIORNI):
                somma = somma_giorni[riep_col]
                ws.cell(row=boll_row, column=boll_col).value = (
                    f"=IF(({somma})>0,ROUND(({somma})*$Q$3,1),0)"
                )
            ws.cell(row=boll_row, column=_LX_MEDIA).value = (
                f"=ROUND(({sums_settimana})*$Q$3/7,1)"
            )

            # Lato destro
            for boll_col, riep_col in zip(_DX_GIORNI, _RIEP_GIORNI):
                somma = somma_giorni[riep_col]
                ws.cell(row=boll_row, column=boll_col).value = (
                    f"=IF(({somma})>0,ROUND(({somma})*$Q$3,1),0)"
                )
            ws.cell(row=boll_row, column=_DX_MEDIA).value = (
                f"=ROUND(({sums_settimana})*$Q$3/7,1)"
            )

    # --- Formattazione condizionale su tutte le righe (famiglie + sottospecie) ---
    for boll_row, assente_max, bassa_max, media_max in BOLL_CF_RIGHE:
        for cf_range in (f"E{boll_row}:L{boll_row}", f"Q{boll_row}:X{boll_row}"):
            ws.conditional_formatting.add(cf_range, CellIsRule(
                operator='greaterThan', formula=[str(media_max)],   fill=CF_ROSSO))
            ws.conditional_formatting.add(cf_range, CellIsRule(
                operator='greaterThan', formula=[str(bassa_max)],   fill=CF_ARANCIO))
            ws.conditional_formatting.add(cf_range, CellIsRule(
                operator='greaterThan', formula=[str(assente_max)], fill=CF_GIALLO))
            ws.conditional_formatting.add(cf_range, CellIsRule(
                operator='greaterThanOrEqual', formula=['0'],        fill=CF_VERDE))


# ---------------------------------------------------------------------------
# Funzione principale
# ---------------------------------------------------------------------------

def applica_formattazione(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb["riepilogo_settimana"]

    # --- Altezza righe ---
    for r in list(range(5, 54)) + list(range(57, 71)):
        ws.row_dimensions[r].height = ALTEZZA_RIGHE

    # --- Sfondo celle di conteggio: verde solo sulle righe rilevanti, bianco sulle altre ---
    # Sinistra: colonne G–M (7–13) | Destra: colonne S–X (19–24)
    # Righe con sfondo verde (pollini di interesse + spore rilevanti):
    VERDE_RIGHE = {10, 17, 18, 19, 20, 22, 23, 24, 27, 28, 29, 38, 40, 41, 52, 58}
    BIANCO = PatternFill(fill_type="solid", fgColor="FFFFFF")

    conteggio_cols_sx = range(7, 14)   # G=7 … M=13
    conteggio_cols_dx = range(19, 25)  # S=19 … X=24
    tutte_righe_dati = list(range(6, 53)) + list(range(58, 70))

    for row in tutte_righe_dati:
        fill = VERDE if row in VERDE_RIGHE else BIANCO
        for col in conteggio_cols_sx:
            ws.cell(row=row, column=col).fill = fill
        for col in conteggio_cols_dx:
            ws.cell(row=row, column=col).fill = fill

    # --- Bordi thin sul perimetro e sulle celle interne ---
    # Sezione sinistra:  D–M  (col 4–13), righe 5–53 e 57–70
    # Sezione destra:    P–X  (col 16–24), righe 5–53 e 57–70
    bordi_sezioni = [
        (range(4, 14), list(range(5, 54)) + list(range(57, 71))),   # D–M
        (range(16, 25), list(range(5, 54)) + list(range(57, 71))),  # P–X
    ]
    for cols, righe in bordi_sezioni:
        for row in righe:
            for col in cols:
                ws.cell(row=row, column=col).border = BORDO_THIN

    # --- Grassetto colonne specie (D e P): solo righe specificate ---
    SPECIE_BOLD_RIGHE = {
        7, 8, 9, 10,
        12, 13, 14, 15, 16, 17, 18,
        21, 22,
        25, 26, 27, 28, 29,
        36, 37, 38, 39, 40,
        44,
        46, 47, 48,
        52, 58,
    }
    specie_cols = [4, 16]  # D e P
    tutte_specie_righe = list(range(6, 53)) + list(range(58, 70))
    for col in specie_cols:
        for row in tutte_specie_righe:
            c = ws.cell(row=row, column=col)
            bold = row in SPECIE_BOLD_RIGHE
            c.font = Font(bold=bold, name=c.font.name, size=c.font.size)

    # --- Grassetto su intestazioni e righe totali (tutte le colonne del range) ---
    # Righe 5 e 57: intestazioni giornate | Righe 53 e 70: totali
    bold_rows_config = {
        5:  list(range(4, 14)) + list(range(16, 25)),  # intestazione pollini
        53: list(range(4, 14)) + list(range(16, 25)),  # totale pollini
        57: list(range(4, 14)) + list(range(16, 25)),  # intestazione spore
        70: list(range(4, 14)) + list(range(16, 25)),  # totale spore
    }
    for row, cols in bold_rows_config.items():
        for col in cols:
            c = ws.cell(row=row, column=col)
            c.font = Font(bold=True, name=c.font.name, size=c.font.size)

    # --- Allineamento centrato su tutto il range dati ---
    # D–M (col 4–13) e P–X (col 16–24), righe 5–53 e 57–70
    allinea_cols = list(range(4, 14)) + list(range(16, 25))
    allinea_righe = list(range(5, 54)) + list(range(57, 71))
    for row in allinea_righe:
        for col in allinea_cols:
            ws.cell(row=row, column=col).alignment = ALIGN_CENTER

    # --- Bollettino: formule famiglia aggregate + formattazione condizionale ---
    aggiorna_bollettino(ws)

    wb.save(path)
    print(f"  Salvato: {path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    targets = [TEMPLATE_PRINCIPALE] + COPIE

    # Backup del template principale
    backup_path = TEMPLATE_PRINCIPALE.with_name(
        TEMPLATE_PRINCIPALE.stem + "_BACKUP.xlsx"
    )
    if not backup_path.exists():
        shutil.copy2(TEMPLATE_PRINCIPALE, backup_path)
        print(f"Backup creato: {backup_path}")
    else:
        print(f"Backup già esistente (non sovrascritto): {backup_path}")

    print("\nApplicazione formattazione...")
    for target in targets:
        if target.exists():
            applica_formattazione(target)
        else:
            print(f"  [SALTATO — non trovato]: {target}")

    print("\nFatto. Aprire il template con LibreOffice/Excel per la verifica visiva.")


if __name__ == "__main__":
    main()
